from fastapi import APIRouter, UploadFile
from openpyxl import load_workbook

a_column = "A2:A1048576"

sheet_router = APIRouter(
  prefix="/sheet",
  tags=["sheet"],
)

@sheet_router.post('')
def apply_functions(file: UploadFile, avg: bool = False):
  wb = load_workbook(file.file)
  ws = wb.active
  row = 2
  if avg:
    ws[f"C{row}"] = "Average"
    ws[f"D{row}"] = f"=Average({a_column})"
    row += 1
  wb.save('test.xlsx')